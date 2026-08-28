"""採点結果を出力する（ローカル xlsx / Google Sheets）。

xlsxとSheetsは `to_dataframe` の同じ列構成を共有しているので、出力先を
増やす場合もこのDataFrameをそのまま書き込むだけでよい。
"""
from __future__ import annotations

import os
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models import ScoredCandidate

COLUMNS = [
    ("rank", "順位"),
    ("total_score", "総合"),
    ("name", "物件名"),
    ("station", "最寄駅"),
    ("walk_min", "徒歩分"),
    ("price_man", "価格(万円)"),
    ("value_score", "割安点"),
    ("asset_score", "資産点"),
    ("commute_score", "通勤点"),
    ("size_score", "広さ点"),
    ("age_score", "築浅点"),
    ("station_score", "駅点"),
    ("hazard_penalty", "ハザ減"),
    ("redflag_penalty", "赤信号減"),
    ("land_sqm", "土地㎡"),
    ("bldg_sqm", "建物㎡"),
    ("layout", "間取り"),
    ("built_year", "築年(西暦)"),
    ("commute_yokohama_min", "通勤_横浜(分)"),
    ("commute_takadanobaba_min", "通勤_高田馬場(分)"),
    ("land_price_used", "採用相場(万/坪)"),
    ("land_price_source", "相場出典"),
    ("estimated_fair_price", "推定適正(万円)"),
    ("corner", "角地"),
    ("hazard_0_3", "ハザード(0-3)"),
    ("redflag_0_3", "赤信号(0-3)"),
    ("zoning", "用途地域"),
    ("address", "住所"),
    ("source_url", "掲載URL"),
    ("intake_source", "取得元"),
    ("note", "メモ"),
    ("id", "ID"),
]


def to_dataframe(scored: list[ScoredCandidate]) -> pd.DataFrame:
    rows = [s.model_dump() for s in scored]
    df = pd.DataFrame(rows)
    keys = [k for k, _ in COLUMNS if k in df.columns]
    df = df[keys]
    df.columns = [dict(COLUMNS)[k] for k in keys]
    return df


def write_xlsx(scored: list[ScoredCandidate], out_path: Path, sheet_name: str = "採点結果") -> Path:
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    df = to_dataframe(scored)

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
        ws = writer.sheets[sheet_name]

        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col_idx in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(
                [len(str(col_name))]
                + [len(str(v)) for v in df[col_name].astype(str).tolist()]
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    return out_path


def write_google_sheet(
    scored: list[ScoredCandidate],
    spreadsheet_id: Optional[str] = None,
    sheet_name: str = "採点結果",
) -> str:
    """採点結果をGoogle Sheetsに書き込む（M5）。

    .env の GOOGLE_SHEETS_SPREADSHEET_ID を使う（未設定なら新規スプレッドシートを作成する）。
    既存シートの内容は毎回まるごと置き換える（冪等性を優先し、追記ではなく上書き）。
    戻り値: スプレッドシートのURL。
    """
    from googleapiclient.discovery import build

    from google_auth import get_credentials

    spreadsheet_id = spreadsheet_id or os.environ.get("GOOGLE_SHEETS_SPREADSHEET_ID")
    creds = get_credentials()
    service = build("sheets", "v4", credentials=creds)

    df = to_dataframe(scored)
    values = [list(df.columns)] + df.astype(object).where(df.notna(), "").values.tolist()

    if not spreadsheet_id:
        created = (
            service.spreadsheets()
            .create(
                body={
                    "properties": {"title": "bukken-pipeline 採点結果"},
                    "sheets": [{"properties": {"title": sheet_name}}],
                }
            )
            .execute()
        )
        spreadsheet_id = created["spreadsheetId"]
        print(
            f"新規スプレッドシートを作成しました。.env の GOOGLE_SHEETS_SPREADSHEET_ID に "
            f"'{spreadsheet_id}' を設定すると次回から同じシートを更新できます。"
        )

    # 既存内容を消してから書き込む（行数が減っても古い行が残らないように）。
    service.spreadsheets().values().clear(
        spreadsheetId=spreadsheet_id, range=sheet_name
    ).execute()
    service.spreadsheets().values().update(
        spreadsheetId=spreadsheet_id,
        range=f"{sheet_name}!A1",
        valueInputOption="RAW",
        body={"values": values},
    ).execute()

    return f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/edit"
